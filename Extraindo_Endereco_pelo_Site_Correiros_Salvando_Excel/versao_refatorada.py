# Importando as bibliotecas necessárias
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time
import os

def pesquisar_endereco_por_cep(cep):
    # Inicializa o WebDriver do Chrome
    abrirNavegador = webdriver.Chrome()
    abrirNavegador.get('https://buscacepinter.correios.com.br/app/endereco/index.php')
    
    # Espera até que o campo de endereço esteja disponível
    time.sleep(3)
    
    # Preenche o campo de endereço com o CEP fornecido
    campo_endereco = abrirNavegador.find_element(By.XPATH, '//*[@id="endereco"]')
    campo_endereco.clear()
    campo_endereco.send_keys(cep)
    
    # Clica no botão de pesquisa
    abrirNavegador.find_element(By.XPATH, '//*[@id="btn_pesquisar"]').click()
    
    # Espera os resultados serem carregados
    time.sleep(4)
    
    # Extrai os dados de rua, bairro, cidade e CEP dos resultados da pesquisa
    rua = abrirNavegador.find_element(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[1]').text
    bairro = abrirNavegador.find_element(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[2]').text
    cidade = abrirNavegador.find_element(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[3]').text
    cep = abrirNavegador.find_element(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[4]').text
    
    # Fecha o navegador
    abrirNavegador.quit()
    
    return rua, bairro, cidade, cep

def main():
    # Caminho do arquivo Excel que contém os CEPs a serem pesquisados e onde serão
    # armazenados os resultados
    nome_arquivo_cep = 'C:\\Valor_do_Dolar_e_Euro\\PesquisaEndereco_3.xlsx'
    
    # Carrega o arquivo Excel e seleciona a planilha 'CEP'
    planilhaDadosEndereco = load_workbook(nome_arquivo_cep)
    sheet_ceps = planilhaDadosEndereco['CEP']
    
    # Limpa a planilha 'Dados' excluindo todas as linhas
    planilhaDadosEndereco.remove(planilhaDadosEndereco['Dados'])
    planilhaDadosEndereco.create_sheet('Dados')
    sheet_dados = planilhaDadosEndereco['Dados']
    
    # Adiciona os cabeçalhos às colunas
    sheet_dados.append(['Endereco', 'Bairro', 'Cidade', 'CEP'])
    
    # Itera sobre as linhas da planilha 'CEP' para pesquisar cada CEP
    for linha in range(2, sheet_ceps.max_row + 1):
        # Obtém o CEP da célula na coluna A da linha atual
        cep = sheet_ceps.cell(row=linha, column=1).value
        
        # Pesquisa o endereço correspondente ao CEP
        rua, bairro, cidade, cep = pesquisar_endereco_por_cep(cep)
        
        # Adiciona os dados do endereço à planilha 'Dados'
        sheet_dados = planilhaDadosEndereco['Dados']
        linha_atual = sheet_dados.max_row + 1
        sheet_dados.cell(row=linha_atual, column=1).value = rua
        sheet_dados.cell(row=linha_atual, column=2).value = bairro
        sheet_dados.cell(row=linha_atual, column=3).value = cidade
        sheet_dados.cell(row=linha_atual, column=4).value = cep
    
    # Salva as alterações no arquivo Excel
    planilhaDadosEndereco.save(filename=nome_arquivo_cep)
    
    # Abre o arquivo Excel após a conclusão das pesquisas
    os.startfile(nome_arquivo_cep)

if __name__ == "__main__":
    main()
