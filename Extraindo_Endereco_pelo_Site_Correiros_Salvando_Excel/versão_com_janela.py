from tkinter import Tk, Label, Entry, Button, messagebox
from tkinter.filedialog import askopenfilename
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time
import os

def pesquisar_endereco_por_cep(cep):
    # Inicializa o WebDriver do Chrome
    abrirNavegador = webdriver.Chrome()
    abrirNavegador.get('https://buscacepinter.correios.com.br/app/endereco/index.php')
    
    # Espera até que o campo de endereço esteja disponível
    time.sleep(3)
    
    # Preenche o campo de endereço com o CEP fornecido
    campo_endereco = abrirNavegador.find_element(By.XPATH, '//*[@id="endereco"]')
    campo_endereco.clear()
    campo_endereco.send_keys(cep)
    
    # Clica no botão de pesquisa
    abrirNavegador.find_element(By.XPATH, '//*[@id="btn_pesquisar"]').click()
    
    # Espera os resultados serem carregados
    time.sleep(4)
    
    # Extrai os dados de rua, bairro, cidade e CEP dos resultados da pesquisa
    rua = abrirNavegador.find_element(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[1]').text
    bairro = abrirNavegador.find_element(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[2]').text
    cidade = abrirNavegador.find_element(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[3]').text
    cep = abrirNavegador.find_element(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[4]').text
    
    # Fecha o navegador
    abrirNavegador.quit()
    
    return rua, bairro, cidade, cep

def processar_pesquisa(nome_arquivo_cep):
    try:
        # Carrega o arquivo Excel e seleciona a planilha 'CEP'
        planilhaDadosEndereco = load_workbook(nome_arquivo_cep)
        sheet_ceps = planilhaDadosEndereco['CEP']
        
        # Limpa a planilha 'Dados' excluindo todas as linhas
        planilhaDadosEndereco.remove(planilhaDadosEndereco['Dados'])
        planilhaDadosEndereco.create_sheet('Dados')
        sheet_dados = planilhaDadosEndereco['Dados']
        
        # Adiciona os cabeçalhos às colunas
        sheet_dados.append(['Endereco', 'Bairro', 'Cidade', 'CEP'])
        
        # Itera sobre as linhas da planilha 'CEP' para pesquisar cada CEP
        for linha in range(2, sheet_ceps.max_row + 1):
            # Obtém o CEP da célula na coluna A da linha atual
            cep = sheet_ceps.cell(row=linha, column=1).value
            
            # Pesquisa o endereço correspondente ao CEP
            rua, bairro, cidade, cep = pesquisar_endereco_por_cep(cep)
            
            # Adiciona os dados do endereço à planilha 'Dados'
            sheet_dados = planilhaDadosEndereco['Dados']
            linha_atual = sheet_dados.max_row + 1
            sheet_dados.cell(row=linha_atual, column=1).value = rua
            sheet_dados.cell(row=linha_atual, column=2).value = bairro
            sheet_dados.cell(row=linha_atual, column=3).value = cidade
            sheet_dados.cell(row=linha_atual, column=4).value = cep
        
        # Salva as alterações no arquivo Excel
        planilhaDadosEndereco.save(filename=nome_arquivo_cep)
        
        # Abre o arquivo Excel após a conclusão das pesquisas
        os.startfile(nome_arquivo_cep)
    except Exception as e:
        messagebox.showerror("Erro", str(e))

def selecionar_arquivo():
    nome_arquivo_cep = askopenfilename(filetypes=(("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")))
    if nome_arquivo_cep:
        entry_nome_arquivo.delete(0, tk.END)
        entry_nome_arquivo.insert(0, nome_arquivo_cep)

def iniciar_pesquisa():
    nome_arquivo_cep = entry_nome_arquivo.get()
    if nome_arquivo_cep:
        processar_pesquisa(nome_arquivo_cep)
    else:
        messagebox.showerror("Erro", "Por favor, selecione um arquivo Excel.")

# Cria a janela principal
root = Tk()
root.title("Pesquisa de Endereço por CEP")
root.geometry("400x150")

# Label e Entry para o nome do arquivo Excel
label_nome_arquivo = Label(root, text="Nome do Arquivo Excel:")
label_nome_arquivo.grid(row=0, column=0, padx=5, pady=5, sticky="w")

entry_nome_arquivo = Entry(root, width=40)
entry_nome_arquivo.grid(row=0, column=1, padx=5, pady=5)

button_selecionar_arquivo = Button(root, text="Selecionar Arquivo", command=selecionar_arquivo)
button_selecionar_arquivo.grid(row=0, column=2, padx=5, pady=5)

# Botão para iniciar a pesquisa
button_iniciar_pesquisa = Button(root, text="Iniciar Pesquisa", command=iniciar_pesquisa)
button_iniciar_pesquisa.grid(row=1, column=1, padx=5, pady=5)

# Loop principal da aplicação
root.mainloop()
