# Importa a função 'load_workbook' do módulo 'openpyxl', que é usada para abrir e manipular arquivos Excel (.xlsx).
from openpyxl import load_workbook

# Importa a classe 'webdriver' do módulo 'selenium' e a renomeia para 'opcoesSelenium'.
# Isso é útil para controlar navegadores de forma programática.
from selenium import webdriver as opcoesSelenium

# Importa a classe 'By' do módulo 'selenium.webdriver.common.by'.
# 'By' é usado para definir os métodos de localização de elementos em uma página web.
from selenium.webdriver.common.by import By

# Importa 'WebDriverWait' para implementar esperas explícitas, aguardando condições específicas.
from selenium.webdriver.support.ui import WebDriverWait

# Importa 'expected_conditions' como 'EC', que contém uma coleção de métodos para esperas explícitas.
from selenium.webdriver.support import expected_conditions as EC

# Define o caminho e o nome do arquivo Excel contendo os dados do formulário.
nomeCaminhoArquivo = 'DadosFormulario.xlsx'

# Abre a planilha Excel especificada.
planilha_aberta = load_workbook(filename=nomeCaminhoArquivo)

# Seleciona a aba específica da planilha chamada 'Dados'.
sheet_selecionada = planilha_aberta['Dados']

# Itera pelas linhas da planilha, começando da segunda linha até a última linha preenchida.
for linha in range(2, len(sheet_selecionada['A']) + 1):
    # Lê os dados de cada célula na linha atual.
    nome = sheet_selecionada[f'A{linha}'].value
    email = sheet_selecionada[f'B{linha}'].value
    telefone = sheet_selecionada[f'C{linha}'].value
    sexo = sheet_selecionada[f'D{linha}'].value 
    sobre = sheet_selecionada[f'E{linha}'].value

    # Inicializa uma instância do navegador Chrome e a atribui à variável 'navegadorFormulario'.
    navegadorFormulario = opcoesSelenium.Chrome()

    # Navega para a URL especificada do formulário online.
    navegadorFormulario.get("https://pt.surveymonkey.com/r/WLXYDX2")

    # Cria uma instância de WebDriverWait para esperar até 10 segundos por condições específicas.
    espera = WebDriverWait(navegadorFormulario, 10)
    
    # Localiza o campo de nome pelo seu atributo 'name' e insere o valor lido do Excel.
    campo_nome = espera.until(EC.presence_of_element_located((By.NAME, "166517069")))
    campo_nome.send_keys(nome)
    
    # Localiza o campo de e-mail pelo seu atributo 'name' e insere o valor lido do Excel.
    campo_email = espera.until(EC.presence_of_element_located((By.NAME, "166517072")))
    campo_email.send_keys(email)

    # Localiza o campo de telefone pelo seu atributo 'name' e insere o valor lido do Excel.
    campo_telefone = espera.until(EC.presence_of_element_located((By.NAME, "166517070")))
    campo_telefone.send_keys(telefone)

    # Localiza o campo 'Sobre' pelo seu atributo 'name' e insere o valor lido do Excel.
    campo_sobre = espera.until(EC.presence_of_element_located((By.NAME, "166517073")))
    campo_sobre.send_keys(sobre)

    # Verifica o valor do campo 'sexo' e seleciona o botão de rádio apropriado.
    if sexo == 'Masculino':
        # Localiza e clica no botão de rádio para 'Masculino' pelo seu ID.
        botao_masculino = espera.until(EC.element_to_be_clickable((By.ID, "166517071_1215509812_label")))
        botao_masculino.click()
    else:
        # Localiza e clica no botão de rádio para 'Feminino' pelo seu ID.
        botao_feminino = espera.until(EC.element_to_be_clickable((By.ID, "166517071_1215509813_label")))
        botao_feminino.click()

    # Localiza o botão de envio do formulário pelo seu XPath.
    enviar = espera.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="patas"]/main/article/section/form/div[2]/button')))

    # Clica no botão de envio do formulário.
    # enviar.click()
