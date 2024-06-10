import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

# Caminho onde estão os arquivos
caminhoArquivos = r"C:\\python_com_rpa_e_criacao_de_projeto_reais\\Excel"

# Variável onde estão todos os arquivos
listaArquivos = os.listdir(caminhoArquivos)

# Listando todos os arquivos + o caminho
listaCaminhoEArquivoExcel = [os.path.join(caminhoArquivos, arquivo) for arquivo in listaArquivos if arquivo.endswith('.xlsx')]

# Criando uma lista para armazenar os DataFrames
dadosList = []

# Copiando todos os dados dos arquivos para a lista de DataFrames
for arquivo in listaCaminhoEArquivoExcel:
    dados = pd.read_excel(arquivo)
    dadosList.append(dados)

# Concatenando todos os DataFrames em um único DataFrame
dadosArquivo = pd.concat(dadosList, ignore_index=True)

# Criando uma nova planilha e passando os dados dos arquivos
dadosArquivo.to_excel(r"C:\\python_com_rpa_e_criacao_de_projeto_reais\\Arquivo Consolidado.xlsx", index=False)

# Carregando a planilha consolidada
caminhoArquivoDadosSistema = r"C:\\python_com_rpa_e_criacao_de_projeto_reais\\Arquivo Consolidado.xlsx"
planilhaDadosSistema = load_workbook(filename=caminhoArquivoDadosSistema)

# Seleciona a sheet1
sheetPlanilhaDadosSistema = planilhaDadosSistema['Sheet1']

# Deleta a coluna A
#sheetPlanilhaDadosSistema.delete_cols(1)

# Renomeia o nome da sheet
sheetPlanilhaDadosSistema.title = 'Dados Consolidados'

# Aumenta a largura das colunas A e B
sheetPlanilhaDadosSistema.column_dimensions['A'].width = 35
sheetPlanilhaDadosSistema.column_dimensions['B'].width = 40

# Criando preenchimento cor Cinza fraco
corCinza = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type='solid')

# Criando preenchimento cor Amarela
corAmarelo = PatternFill(start_color='00FFFFCC', end_color='00FFFFCC', fill_type='solid')

# Coloca borda preta na célula
bfFina = Side(border_style='thin', color='000000')
borda = Border(left=bfFina, right=bfFina, top=bfFina, bottom=bfFina)

# Colocando cores e borda e fundo nas celulas A1, B1 e C1
sheetPlanilhaDadosSistema['A1'].fill = corAmarelo
sheetPlanilhaDadosSistema['B1'].fill = corAmarelo
sheetPlanilhaDadosSistema['C1'].fill = corAmarelo
sheetPlanilhaDadosSistema['A1'].border = borda
sheetPlanilhaDadosSistema['B1'].border = borda
sheetPlanilhaDadosSistema['C1'].border = borda

# Dando um for linha por linha para pintar as bordas e o fundo
for linha in range(2, len(sheetPlanilhaDadosSistema['A']) + 1):
    # Criando celula A2, B2, C2 e assim por diante de acordo com o número da linha
    celulaColunaA = 'A' + str(linha)
    celulaColunaB = 'B' + str(linha)
    celulaColunaC = 'C' + str(linha)
    
    # Colorindo o fundo de cinza claro das linhas
    sheetPlanilhaDadosSistema[celulaColunaA].fill = corCinza
    sheetPlanilhaDadosSistema[celulaColunaB].fill = corCinza
    sheetPlanilhaDadosSistema[celulaColunaC].fill = corCinza
    
    # Colocando a borda preta nas linhas
    sheetPlanilhaDadosSistema[celulaColunaA].border = borda
    sheetPlanilhaDadosSistema[celulaColunaB].border = borda
    sheetPlanilhaDadosSistema[celulaColunaC].border = borda

# Achando a ultima linha e somando + 1
ultimaLinhaMaisUm = linha + 1
print(ultimaLinhaMaisUm)

# Desenhando a coluna C + a última linha para usar na variável
celulaUltimaLinha = 'C' + str(ultimaLinhaMaisUm)

# Desenhando a fórmula da soma
formulaSoma = "=SUM(C2:C" + str(linha) + ")"

# Imprimindo a fórmula de soma na planilha
sheetPlanilhaDadosSistema[celulaUltimaLinha] = formulaSoma

# Salva a planilha com as alterações
planilhaDadosSistema.save(filename=caminhoArquivoDadosSistema)

# Abre a planilha
os.startfile(caminhoArquivoDadosSistema)
