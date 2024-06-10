# Importa a função load_workbook da biblioteca openpyxl para manipular arquivos Excel
from openpyxl import load_workbook
# Importa a biblioteca os para interagir com o sistema operacional (não usada neste trecho de código, mas pode ser útil)
import os

# Importa a biblioteca win32com.client, que permite automação de tarefas do Windows, como interagir com o Outlook
import win32com.client as win32

# Cria um objeto que representa o aplicativo Outlook
outlook = win32.Dispatch('outlook.application')

# Nome do arquivo Excel que contém a lista de emails
nome_arquivo = "ListaEmail.xlsx"
# Abre o arquivo Excel
planilha_aberta = load_workbook(filename=nome_arquivo)

# Seleciona a planilha chamada 'Dados'
sheet_selecionada = planilha_aberta['Dados']

# Loop que percorre todas as linhas da planilha, começando da segunda linha
for linha in range(2, len(sheet_selecionada['A']) + 1):
    
    # Lê o valor da coluna A na linha atual (nome)
    nome = sheet_selecionada['A%s' % linha].value
    # Lê o valor da coluna B na linha atual (nome completo)
    nomeCompleto = sheet_selecionada['B%s' % linha].value
    # Lê o valor da coluna C na linha atual (email)
    email = sheet_selecionada['C%s' % linha].value
    
    # Cria um novo item de email no Outlook
    emailOutlook = outlook.CreateItem(0)

    # Define o destinatário do email
    emailOutlook.To = email
    # Define o assunto do email, incluindo o nome completo
    emailOutlook.Subject = "Lista de vendas " + nomeCompleto
    # Define o corpo do email em formato HTML, incluindo o nome e uma imagem
    emailOutlook.HTMLBody = f"""
    <p>Boa noite <b>{nome}</b>.</p>
    <p>Segue o relatório com suas vendas.</p>
    <p>Atenciosamente.</p>
    <p><img src="C:\\python_com_rpa_e_criacao_de_projeto_reais\\tech.jpg"></p>
    """
    
    # Define o caminho do arquivo de anexo baseado no nome completo
    anexoEmail = "C:\\python_com_rpa_e_criacao_de_projeto_reais\\" + nomeCompleto + ".xlsx"
    # Adiciona o arquivo como anexo ao email
    emailOutlook.Attachments.Add(anexoEmail)

    # Salva o email no Outlook sem enviar
    emailOutlook.save() # use emailOutlook.Send() para enviar o email
