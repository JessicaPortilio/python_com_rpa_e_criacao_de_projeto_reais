# Importa as funções load_workbook e Workbook da biblioteca openpyxl
from openpyxl import load_workbook, Workbook  
# Importa PatternFill para aplicar cores nas células
from openpyxl.styles import PatternFill  
import os  # Importa a biblioteca os para operações de sistema

# Define o nome do arquivo Excel a ser modificado ou criado
caminho_nome_arquivo = "InserirDadosPintarCelulas.xlsx"

# Verifica se o arquivo existe
if os.path.exists(caminho_nome_arquivo):
    # Carrega o arquivo existente
    planilha_aberta = load_workbook(filename=caminho_nome_arquivo)
else:
    # Cria um novo workbook e uma planilha chamada "Professor"
    planilha_aberta = Workbook()
    planilha_aberta.create_sheet(title="Professor", index=0)
    # Remove a planilha padrão criada automaticamente (chamada "Sheet")
    if "Sheet" in planilha_aberta.sheetnames:
        std_sheet = planilha_aberta["Sheet"]
        planilha_aberta.remove(std_sheet)

# Seleciona a planilha chamada "Professor"
sheet_selecionada = planilha_aberta['Professor']

# Dados a serem inseridos na planilha
dadosTabela = [
    ['Nome', 'Idade'],
    ['Berenice', 28],
    ['Caio', 32],
    ['Nicole', 34],
    ['Leonardo', 19],
    ['Amanda', 25]
]

# Insere os dados na planilha
for linhaPlanilha in dadosTabela:
    sheet_selecionada.append(linhaPlanilha)

# Define a cor de fundo das células do título (primeira linha)
corTitulo = PatternFill(start_color='00FFFF00',  # Cor amarela
                      end_color='00FFFF00',
                      fill_type='solid')

# Define a cor de fundo das células de dados (demais linhas)
corCelulas = PatternFill(start_color='00CCFFCC',  # Cor verde clara
                      end_color='00CCFFCC',
                      fill_type='solid')

# Aplica a cor de fundo nas células do título
sheet_selecionada["A1"].fill = corTitulo
sheet_selecionada["B1"].fill = corTitulo

# Aplica a cor de fundo nas células de dados
for linha in range(2, len(sheet_selecionada['A']) + 1):
    celulaColunaA = "A" + str(linha)
    celulaColunaB = "B" + str(linha)
    sheet_selecionada[celulaColunaA].fill = corCelulas
    sheet_selecionada[celulaColunaB].fill = corCelulas

# Salva o arquivo Excel com as alterações realizadas
planilha_aberta.save(filename=caminho_nome_arquivo)

# Abre o arquivo Excel modificado
os.startfile(caminho_nome_arquivo)
