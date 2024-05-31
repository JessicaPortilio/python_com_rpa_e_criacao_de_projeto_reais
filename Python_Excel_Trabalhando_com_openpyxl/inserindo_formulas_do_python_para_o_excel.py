# Importa a função 'load_workbook' do módulo 'openpyxl', que permite carregar arquivos Excel.
from openpyxl import load_workbook

# Importa o módulo 'os', que fornece funções para interagir com o sistema operacional.
import os

# Define o caminho e nome do arquivo Excel que será aberto.
caminho_nome_arquivo = "Formulas2.xlsx"

# Carrega a planilha Excel especificada pelo caminho e nome do arquivo.
planilha_aberta = load_workbook(filename=caminho_nome_arquivo)

# Seleciona a planilha específica chamada 'Aluno' dentro do arquivo Excel.
sheet_selecionada = planilha_aberta['Aluno']

# Preenche as células com fórmulas de soma nas colunas A e B.
sheet_selecionada['A6'] = "=SUM(A2:A5)"  # Soma os valores de A2 a A5 e coloca o resultado em A6.
sheet_selecionada['B6'] = "=SUM(B2:B5)"  # Soma os valores de B2 a B5 e coloca o resultado em B6.

# Preenche as células com fórmulas de operações aritméticas na coluna D.
sheet_selecionada['D2'] = "=A2+B2"  # Soma os valores das células A2 e B2 e coloca o resultado em D2.
sheet_selecionada['D3'] = "=A3-B3"  # Subtrai o valor da célula B3 de A3 e coloca o resultado em D3.
sheet_selecionada['D4'] = "=A4*B4"  # Multiplica os valores das células A4 e B4 e coloca o resultado em D4.
sheet_selecionada['D5'] = "=A5/B5"  # Divide o valor da célula A5 pelo valor de B5 e coloca o resultado em D5.

# Preenche as células com fórmulas de MID para separar partes de um CPF.
sheet_selecionada['B12'] = "=MID(A12,1,3)"  # Extrai os primeiros 3 caracteres da célula A12 e coloca o resultado em B12.
sheet_selecionada['C12'] = "=MID(A12,5,3)"  # Extrai 3 caracteres começando do 5º da célula A12 e coloca o resultado em C12.
sheet_selecionada['D12'] = "=MID(A12,9,3)"  # Extrai 3 caracteres começando do 9º da célula A12 e coloca o resultado em D12.
sheet_selecionada['E12'] = "=MID(A12,13,2)"  # Extrai 2 caracteres começando do 13º da célula A12 e coloca o resultado em E12.

# Salva a planilha Excel com as alterações feitas.
planilha_aberta.save(filename=caminho_nome_arquivo)

# Abre o arquivo Excel usando o programa padrão associado no sistema operacional.
os.startfile(caminho_nome_arquivo)
