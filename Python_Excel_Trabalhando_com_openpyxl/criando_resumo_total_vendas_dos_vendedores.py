# Importa a função 'load_workbook' do módulo 'openpyxl', que permite carregar arquivos Excel.
from openpyxl import load_workbook

# Importa o módulo 'os', que fornece funções para interagir com o sistema operacional.
import os

# Define o caminho e nome do arquivo Excel que será aberto.
caminho_nome_arquivo = "Vendedores.xlsx"

# Carrega a planilha Excel especificada pelo caminho e nome do arquivo.
planilha_aberta = load_workbook(filename=caminho_nome_arquivo)

# Seleciona a planilha específica chamada 'Vendas' dentro do arquivo Excel.
sheet_selecionada = planilha_aberta['Vendas']

# Inicializa variáveis para somar as vendas de cada vendedor.
somarAmandaMartins = 0
somarElianeMoreira = 0
somarLeonardoAlmeida = 0
somarNicolasPereira = 0

# Itera pelas linhas da planilha, começando da segunda linha até a última.
for linha in range(2, len(sheet_selecionada['A']) + 1):
    # Verifica se o valor na coluna A da linha atual corresponde a "Amanda Martins".
    if sheet_selecionada['A%s' % linha].value == "Amanda Martins":
        # Adiciona o valor da coluna C da linha atual à soma de vendas de Amanda Martins.
        somarAmandaMartins += sheet_selecionada['C%s' % linha].value
        
    # Verifica se o valor na coluna A da linha atual corresponde a "Eliane Moreira".
    elif sheet_selecionada['A%s' % linha].value == "Eliane Moreira":
        # Adiciona o valor da coluna C da linha atual à soma de vendas de Eliane Moreira.
        somarElianeMoreira += sheet_selecionada['C%s' % linha].value
        
    # Verifica se o valor na coluna A da linha atual corresponde a "Nicolas Pereira".
    elif sheet_selecionada['A%s' % linha].value == "Nicolas Pereira":
        # Adiciona o valor da coluna C da linha atual à soma de vendas de Nicolas Pereira.
        somarNicolasPereira += sheet_selecionada['C%s' % linha].value
        
    # Verifica se o valor na coluna A da linha atual corresponde a "Leonardo Almeida".
    elif sheet_selecionada['A%s' % linha].value == "Leonardo Almeida":
        # Adiciona o valor da coluna C da linha atual à soma de vendas de Leonardo Almeida.
        somarLeonardoAlmeida += sheet_selecionada['C%s' % linha].value

# Cria uma nova planilha chamada "Resumo" para armazenar o resumo das vendas.
sheet_resumo = planilha_aberta.create_sheet(title="Resumo")

# Preenche os títulos das colunas na nova planilha "Resumo".
sheet_resumo['A1'] = "Vendedores"
sheet_resumo['B1'] = "Vendas"

# Preenche os dados do primeiro vendedor (Amanda Martins) e suas vendas totais.
sheet_resumo['A2'] = "Amanda Martins"
sheet_resumo['B2'] = somarAmandaMartins

# Preenche os dados do segundo vendedor (Eliane Moreira) e suas vendas totais.
sheet_resumo['A3'] = "Eliane Moreira"
sheet_resumo['B3'] = somarElianeMoreira

# Preenche os dados do terceiro vendedor (Leonardo Almeida) e suas vendas totais.
sheet_resumo['A4'] = "Leonardo Almeida"
sheet_resumo['B4'] = somarLeonardoAlmeida

# Preenche os dados do quarto vendedor (Nicolas Pereira) e suas vendas totais.
sheet_resumo['A5'] = "Nicolas Pereira"
sheet_resumo['B5'] = somarNicolasPereira

# Salva a planilha com as alterações realizadas.
planilha_aberta.save(filename=caminho_nome_arquivo)

# Abre a planilha com as alterações no sistema operacional.
os.startfile(caminho_nome_arquivo)
