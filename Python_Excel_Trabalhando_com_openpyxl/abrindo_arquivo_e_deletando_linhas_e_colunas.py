# Importa a função load_workbook da biblioteca openpyxl
from openpyxl import load_workbook  
# Importa a biblioteca os para operações de sistema
import os  

# Define o nome do arquivo Excel a ser modificado
caminho_nome_arquivo = "DeletarLinhaColuna.xlsx"
# Carrega o arquivo Excel
planilha_aberta = load_workbook(filename=caminho_nome_arquivo)

# Seleciona a planilha chamada "Aluno"
sheet_selecionada = planilha_aberta['Aluno']

# Deleta a terceira linha (originalmente) da planilha
sheet_selecionada.delete_rows(3)
# Deleta novamente a terceira linha (a quarta linha original)
sheet_selecionada.delete_rows(3)
# Deleta a quinta linha (a sexta linha original)
sheet_selecionada.delete_rows(5)

# Deleta a coluna B (segunda coluna)
sheet_selecionada.delete_cols(2)

# Salva o arquivo Excel com as alterações realizadas
planilha_aberta.save(filename=caminho_nome_arquivo)

# Abre o arquivo Excel modificado
os.startfile(caminho_nome_arquivo)
