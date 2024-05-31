# Importa as funções 'load_workbook' e 'Workbook' do módulo 'openpyxl', que permitem carregar e criar arquivos Excel.
from openpyxl import load_workbook
from openpyxl import Workbook

# Importa o módulo 'os', que fornece funções para interagir com o sistema operacional.
import os

# Define o caminho e nome do arquivo Excel que contém os dados do sistema.
caminhoArquivoDadosSistema = "DadosSistema.xlsx"

# Carrega a planilha Excel especificada pelo caminho e nome do arquivo.
planilhaDadosSistema = load_workbook(filename=caminhoArquivoDadosSistema)

# Seleciona a planilha específica chamada 'Dados' dentro do arquivo Excel.
sheetPlanilhaDadosSistema = planilhaDadosSistema['Dados']

# Cria um novo arquivo Excel.
criandoNovoArquivoExcel = Workbook()

# Seleciona a planilha ativa no novo arquivo Excel.
nova_planilha = criandoNovoArquivoExcel.active

# Copia os dados da planilha 'Dados' para a nova planilha.
for linha in range(1, len(sheetPlanilhaDadosSistema['A']) + 1):
    for coluna in range(1, 10):
        nova_planilha.cell(row=linha, column=coluna).value = sheetPlanilhaDadosSistema.cell(row=linha, column=coluna).value

# Deleta a segunda linha da nova planilha.
nova_planilha.delete_rows(2)

# Deleta a segunda e terceira colunas da nova planilha (a segunda coluna é deletada duas vezes para remover duas colunas adjacentes).
nova_planilha.delete_cols(2)
nova_planilha.delete_cols(2)

# Renomeia a planilha ativa para 'Dados Funcionarios'.
nova_planilha.title = 'Dados Funcionarios'

# Cria uma nova planilha e renomeia como 'Resumo'.
criandoNovoArquivoExcel.create_sheet('Resumo')

# Seleciona a planilha 'Resumo' no novo arquivo Excel.
selecionaSheetResumoNovaPlanilha = criandoNovoArquivoExcel['Resumo']

# Preenche os títulos das colunas na planilha 'Resumo'.
selecionaSheetResumoNovaPlanilha['A1'] = "Vendedor"
selecionaSheetResumoNovaPlanilha['B1'] = "Total Vendas"

# Preenche os dados do primeiro vendedor e usa a fórmula SUMIF para calcular o total de vendas.
selecionaSheetResumoNovaPlanilha['A2'] = "Amanda Martins"
selecionaSheetResumoNovaPlanilha['B2'] = "=SUMIF('Dados Funcionarios'!A:C,A2,'Dados Funcionarios'!C:C)"

# Preenche os dados do segundo vendedor e usa a fórmula SUMIF para calcular o total de vendas.
selecionaSheetResumoNovaPlanilha['A3'] = "Eliane Moreira"
selecionaSheetResumoNovaPlanilha['B3'] = "=SUMIF('Dados Funcionarios'!A:C,A3,'Dados Funcionarios'!C:C)"

# Preenche os dados do terceiro vendedor e usa a fórmula SUMIF para calcular o total de vendas.
selecionaSheetResumoNovaPlanilha['A4'] = "Leonardo Almeida"
selecionaSheetResumoNovaPlanilha['B4'] = "=SUMIF('Dados Funcionarios'!A:C,A4,'Dados Funcionarios'!C:C)"

# Preenche os dados do quarto vendedor e usa a fórmula SUMIF para calcular o total de vendas.
selecionaSheetResumoNovaPlanilha['A5'] = "Nicolas Pereira"
selecionaSheetResumoNovaPlanilha['B5'] = "=SUMIF('Dados Funcionarios'!A:C,A5,'Dados Funcionarios'!C:C)"

# Define o caminho e nome para o novo arquivo Excel que será salvo.
caminhoNovaPlanilha = "RelatorioSistema.xlsx"

# Salva a planilha com as alterações no novo arquivo Excel.
criandoNovoArquivoExcel.save(filename=caminhoNovaPlanilha)

# Abre o novo arquivo Excel no sistema operacional.
os.startfile(caminhoNovaPlanilha)
