from openpyxl import load_workbook, Workbook
import os

caminho_nome_arquivo = "InserirDados.xlsx"

# Verifica se o arquivo existe
if os.path.exists(caminho_nome_arquivo):
    # Carrega o arquivo existente
    planilha_aberta = load_workbook(filename=caminho_nome_arquivo)
else:
    # Cria um novo workbook e uma planilha chamada "Aluno"
    planilha_aberta = Workbook()
    planilha_aberta.create_sheet(title="Aluno", index=0)
    # Remove a planilha padrão criada automaticamente (chamada "Sheet")
    if "Sheet" in planilha_aberta.sheetnames:
        std_sheet = planilha_aberta["Sheet"]
        planilha_aberta.remove(std_sheet)

#Seleciona a sheet de Professor
sheet_selecionada = planilha_aberta['Aluno']

#Popula as informações que vão para a planilha
dadosTabela = [
    ['Nome', 'Idade'],
    ['Berenice', 28],
    ['Caio', 32],
    ['Nicole', 34],
    ['Leonardo', 19],
    ['Amanda', 25]
]

# O append pega toda a lista e passa para a sheet
for linhaPlanilha in dadosTabela:
    sheet_selecionada.append(linhaPlanilha)



#Salva a planilha com as alterações
planilha_aberta.save(filename=caminho_nome_arquivo)

#Abre a planilha
os.startfile(caminho_nome_arquivo)