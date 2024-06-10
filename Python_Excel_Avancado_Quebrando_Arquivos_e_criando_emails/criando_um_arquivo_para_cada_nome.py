# Importa a função 'load_workbook' do módulo 'openpyxl' para carregar arquivos Excel.
from openpyxl import load_workbook
# Importa a classe 'Workbook' do módulo 'openpyxl' para criar novos arquivos Excel.
from openpyxl import Workbook
# Importa o módulo 'os' que fornece funções para interagir com o sistema operacional.
import os

# Define o caminho e nome do arquivo Excel que será processado.
nome_arquivo = "Quebrar2.xlsx"

# Carrega a planilha Excel especificada pelo caminho e nome do arquivo.
planilha_aberta = load_workbook(filename=nome_arquivo)

# Seleciona a planilha específica chamada 'Dados' dentro do arquivo Excel.
sheet_selecionada = planilha_aberta['Dados']

# Cria um novo arquivo Excel para armazenar as planilhas separadas.
criandoNovoArquivoExcel = Workbook()

# Variável para armazenar o nome do funcionário atual, inicializada como string vazia.
nomeNovo = ""

# Calcula o total de linhas na coluna 'A' da planilha 'Dados'.
totalLinha = len(sheet_selecionada['A']) + 1

# Itera sobre cada linha da planilha, começando da segunda linha.
for linha in range(2, len(sheet_selecionada['A']) + 1):
    
    # Armazena o nome do funcionário da linha atual.
    nomeAtual = sheet_selecionada['A%s' % linha].value
    
    # Verifica se o nome do funcionário atual é o mesmo do anterior.
    if nomeNovo == nomeAtual:
        # Determina a próxima linha vazia na planilha do funcionário atual.
        linhaSheetQuebra = len(selecionaSheetVendasNovaPlanilha['A']) + 1
        celulaColunaA = "A" + str(linhaSheetQuebra)
        celulaColunaB = "B" + str(linhaSheetQuebra)
        celulaColunaC = "C" + str(linhaSheetQuebra)
        
        # Preenche os dados na nova planilha do funcionário atual.
        selecionaSheetVendasNovaPlanilha[celulaColunaA] = sheet_selecionada['A%s' % linha].value
        selecionaSheetVendasNovaPlanilha[celulaColunaB] = sheet_selecionada['B%s' % linha].value
        selecionaSheetVendasNovaPlanilha[celulaColunaC] = sheet_selecionada['C%s' % linha].value
        
        # Salva o arquivo Excel após adicionar os dados do funcionário.
        criandoNovoArquivoExcel.save(filename=caminhoNovaPlanilha)
        
    else:
        # Atualiza a variável nomeNovo com o nome do funcionário atual.
        nomeNovo = sheet_selecionada['A%s' % linha].value
        
        # Define a planilha ativa como a nova planilha que será criada.
        nova_planilha = criandoNovoArquivoExcel.active
        
        # Define o título da nova planilha como "Vendas".
        nova_planilha.title = "Vendas"
        
        # Define o caminho e nome do novo arquivo Excel baseado no nome do funcionário.
        caminhoNovaPlanilha = sheet_selecionada['A%s' % linha].value + ".xlsx"
        
        # Seleciona a nova planilha de vendas criada.
        selecionaSheetVendasNovaPlanilha = criandoNovoArquivoExcel['Vendas']
        
        # Adiciona os títulos às colunas na nova planilha.
        selecionaSheetVendasNovaPlanilha['A1'] = "Vendedor"
        selecionaSheetVendasNovaPlanilha['B1'] = "Produtos"
        selecionaSheetVendasNovaPlanilha['C1'] = "Vendas"
        
        # Preenche as informações do funcionário na segunda linha da nova planilha.
        selecionaSheetVendasNovaPlanilha['A2'] = sheet_selecionada['A%s' % linha].value
        selecionaSheetVendasNovaPlanilha['B2'] = sheet_selecionada['B%s' % linha].value
        selecionaSheetVendasNovaPlanilha['C2'] = sheet_selecionada['C%s' % linha].value
        
        # Deleta linhas desnecessárias da planilha de vendas para manter o arquivo limpo.
        selecionaSheetVendasNovaPlanilha.delete_rows(3, 100)
        
        # Salva o novo arquivo Excel após adicionar os dados do funcionário.
        criandoNovoArquivoExcel.save(filename=caminhoNovaPlanilha)
