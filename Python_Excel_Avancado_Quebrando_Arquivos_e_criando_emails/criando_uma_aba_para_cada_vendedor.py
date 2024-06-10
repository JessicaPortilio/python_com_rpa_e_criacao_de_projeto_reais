# Importa a função 'load_workbook' do módulo 'openpyxl' para carregar arquivos Excel.
from openpyxl import load_workbook

# Importa o módulo 'os' que fornece funções para interagir com o sistema operacional.
import os

# Define o caminho e nome do arquivo Excel que será processado.
nome_arquivo = "Quebrar1.xlsx"

# Carrega a planilha Excel especificada pelo caminho e nome do arquivo.
planilha_aberta = load_workbook(filename=nome_arquivo)

# Seleciona a planilha específica chamada 'Dados' dentro do arquivo Excel.
sheet_selecionada = planilha_aberta['Dados']

# Variável para armazenar o nome do funcionário atual, inicializada como string vazia.
nomeNovo = ""

# Calcula o total de linhas na coluna 'A' da planilha 'Dados'.
totalLinha = len(sheet_selecionada['A']) + 1

# Itera sobre cada linha da planilha, começando da segunda linha.
for linha in range(2, len(sheet_selecionada['A']) + 1):
    
    # Armazena o nome do funcionário da linha atual.
    nomeAtual = sheet_selecionada['A%s' % linha].value
    
    # Verifica se o nome do funcionário atual é o mesmo do anterior.
    if nomeNovo == nomeAtual:
        # Determina a próxima linha vazia na planilha do funcionário atual.
        linhaSheetQuebra = len(sheet_selecionada2['A']) + 1
        celulaColunaA = "A" + str(linhaSheetQuebra)
        celulaColunaB = "B" + str(linhaSheetQuebra)
        celulaColunaC = "C" + str(linhaSheetQuebra)
        
        # Preenche os dados na nova planilha do funcionário atual.
        sheet_selecionada2[celulaColunaA] = sheet_selecionada['A%s' % linha].value
        sheet_selecionada2[celulaColunaB] = sheet_selecionada['B%s' % linha].value
        sheet_selecionada2[celulaColunaC] = sheet_selecionada['C%s' % linha].value
        
    else:
        # Cria uma nova planilha com o nome do funcionário atual.
        sheet_resumo = planilha_aberta.create_sheet(title=nomeAtual)
        
        # Seleciona a nova planilha criada.
        sheet_selecionada2 = planilha_aberta[nomeAtual]
        
        # Atualiza a variável nomeNovo com o nome do funcionário atual.
        nomeNovo = sheet_selecionada['A%s' % linha].value
        
        # Adiciona os títulos às colunas na nova planilha.
        sheet_selecionada2['A1'] = "Vendedor"
        sheet_selecionada2['B1'] = "Produtos"
        sheet_selecionada2['C1'] = "Vendas"
        
        # Preenche as informações do funcionário na segunda linha da nova planilha.
        sheet_selecionada2['A2'] = sheet_selecionada['A%s' % linha].value
        sheet_selecionada2['B2'] = sheet_selecionada['B%s' % linha].value
        sheet_selecionada2['C2'] = sheet_selecionada['C%s' % linha].value
        
# Salva as alterações feitas na planilha original.
planilha_aberta.save(filename=nome_arquivo)

# Abre o arquivo Excel com as mudanças feitas.
os.startfile(nome_arquivo)
