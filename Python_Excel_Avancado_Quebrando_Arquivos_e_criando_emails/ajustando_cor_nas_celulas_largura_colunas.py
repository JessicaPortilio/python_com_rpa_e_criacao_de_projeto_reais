from openpyxl import load_workbook
from openpyxl import Workbook
import os

# Importa estilos específicos do módulo 'openpyxl' para formatar células.
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles import colors
from openpyxl.cell import Cell

# Define o caminho e nome do arquivo Excel que será processado.
nome_arquivo = "Quebrar2.xlsx"
# Carrega a planilha Excel especificada pelo caminho e nome do arquivo.
planilha_aberta = load_workbook(filename=nome_arquivo)

# Seleciona a planilha específica chamada 'Dados' dentro do arquivo Excel.
sheet_selecionada = planilha_aberta['Dados']

# Cria um novo arquivo Excel para armazenar as planilhas separadas.
criandoNovoArquivoExcel = Workbook()

# Variável para armazenar o nome do funcionário atual, inicializada como string vazia.
nomeNovo = ""
# Calcula o total de linhas na coluna 'A' da planilha 'Dados'.
totalLinha = len(sheet_selecionada['A']) + 1

# Cria um preenchimento de cor azul para as células.
corAzul = PatternFill(start_color='0099CCFF', end_color='0099CCFF', fill_type='solid')
# Cria um preenchimento de cor amarela para as células.
corAmarelo = PatternFill(start_color='00FFFFCC', end_color='00FFFFCC', fill_type='solid')

# Define a borda fina preta para as células.
bdFina = Side(border_style="thin", color="000000")
borda = Border(left=bdFina, right=bdFina, top=bdFina, bottom=bdFina)

# Itera sobre cada linha da planilha, começando da segunda linha.
for linha in range(2, len(sheet_selecionada['A']) + 1):
    # Armazena o nome do funcionário da linha atual.
    nomeAtual = sheet_selecionada['A%s' % linha].value
    
    # Verifica se o nome do funcionário atual é o mesmo do anterior.
    if nomeNovo == nomeAtual:
        # Determina a próxima linha vazia na planilha do funcionário atual.
        linhaSheetQuebra = len(selecionaSheetVendasNovaPlanilha['A']) + 1
        celulaColunaA = "A" + str(linhaSheetQuebra)
        celulaColunaB = "B" + str(linhaSheetQuebra)
        celulaColunaC = "C" + str(linhaSheetQuebra)
        
        # Preenche os dados na nova planilha do funcionário atual.
        selecionaSheetVendasNovaPlanilha[celulaColunaA] = sheet_selecionada['A%s' % linha].value
        selecionaSheetVendasNovaPlanilha[celulaColunaB] = sheet_selecionada['B%s' % linha].value
        selecionaSheetVendasNovaPlanilha[celulaColunaC] = sheet_selecionada['C%s' % linha].value
        
        # Aplica a cor amarela às células da linha adicionada.
        selecionaSheetVendasNovaPlanilha[celulaColunaA].fill = corAmarelo
        selecionaSheetVendasNovaPlanilha[celulaColunaB].fill = corAmarelo
        selecionaSheetVendasNovaPlanilha[celulaColunaC].fill = corAmarelo
        
        # Salva o arquivo Excel após adicionar os dados do funcionário.
        criandoNovoArquivoExcel.save(filename=caminhoNovaPlanilha)
        
    else:
        # Atualiza a variável nomeNovo com o nome do funcionário atual.
        nomeNovo = sheet_selecionada['A%s' % linha].value
        
        # Define a planilha ativa como a nova planilha que será criada.
        nova_planilha = criandoNovoArquivoExcel.active
        # Define o título da nova planilha como "Vendas".
        nova_planilha.title = "Vendas"
        
        # Define o caminho e nome do novo arquivo Excel baseado no nome do funcionário.
        caminhoNovaPlanilha = sheet_selecionada['A%s' % linha].value + ".xlsx"
        
        # Seleciona a nova planilha de vendas criada.
        selecionaSheetVendasNovaPlanilha = criandoNovoArquivoExcel['Vendas']
        
        # Adiciona os títulos às colunas na nova planilha.
        selecionaSheetVendasNovaPlanilha['A1'] = "Vendedor"
        selecionaSheetVendasNovaPlanilha['B1'] = "Produtos"
        selecionaSheetVendasNovaPlanilha['C1'] = "Vendas"
        
        # Preenche as informações do funcionário na segunda linha da nova planilha.
        selecionaSheetVendasNovaPlanilha['A2'] = sheet_selecionada['A%s' % linha].value
        selecionaSheetVendasNovaPlanilha['B2'] = sheet_selecionada['B%s' % linha].value
        selecionaSheetVendasNovaPlanilha['C2'] = sheet_selecionada['C%s' % linha].value
        
        # Aplica a cor azul e a borda às células do cabeçalho.
        selecionaSheetVendasNovaPlanilha['A1'].fill = corAzul
        selecionaSheetVendasNovaPlanilha['B1'].fill = corAzul
        selecionaSheetVendasNovaPlanilha['C1'].fill = corAzul
        selecionaSheetVendasNovaPlanilha['A1'].border = borda
        selecionaSheetVendasNovaPlanilha['B1'].border = borda
        selecionaSheetVendasNovaPlanilha['C1'].border = borda
        
        # Aplica a cor amarela às células da segunda linha.
        selecionaSheetVendasNovaPlanilha['A2'].fill = corAmarelo
        selecionaSheetVendasNovaPlanilha['B2'].fill = corAmarelo
        selecionaSheetVendasNovaPlanilha['C2'].fill = corAmarelo
        
        # Aumenta o tamanho das colunas para melhor visualização.
        selecionaSheetVendasNovaPlanilha.column_dimensions["A"].width = 18
        selecionaSheetVendasNovaPlanilha.column_dimensions["B"].width = 25
        selecionaSheetVendasNovaPlanilha.column_dimensions["C"].width = 15
        
        # Remove linhas extras da planilha para manter o arquivo limpo.
        selecionaSheetVendasNovaPlanilha.delete_rows(3, 100)
        
        # Salva o novo arquivo Excel após adicionar os dados do funcionário.
        criandoNovoArquivoExcel.save(filename=caminhoNovaPlanilha)